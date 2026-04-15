import os
import pandas as pd
from typing import Optional, Dict, List, Any
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelDB:
    def __init__(self, filename="finanzas.xlsx"):
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        self.filename = os.path.join(BASE_DIR, filename)
        # Rediseño: Cada categoría tiene su propia hoja
        self.sheets = ["Dashboard", "Ingresos", "Gastos", "Ahorros"]
        self.columns = ["Fecha", "Categoria", "Descripcion", "Monto"]
        self._init_db()

    def _init_db(self) -> None:
        """Creates the Excel file with the required sheets if it doesn't exist."""
        if not os.path.exists(self.filename):
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                # Crear el Dashboard vacío pero con una estructura básica
                df_dash = pd.DataFrame({"Resumen": ["Total Ingresos", "Total Gastos", "Total Ahorros", "Balance Neto"], "Monto": [0, 0, 0, 0]})
                df_dash.to_excel(writer, index=False, sheet_name="Dashboard")
                
                # Crear las hojas individuales
                df_empty = pd.DataFrame(columns=self.columns)
                df_empty.to_excel(writer, index=False, sheet_name="Ingresos")
                df_empty.to_excel(writer, index=False, sheet_name="Gastos")
                df_empty.to_excel(writer, index=False, sheet_name="Ahorros")
            
            self._format_excel()
        else:
            # Archivo existe, pero podría ser viejo (ej. solo hoja "Registros")
            try:
                wb = load_workbook(self.filename)
                needs_save = False
                
                if "Dashboard" not in wb.sheetnames:
                    ws = wb.create_sheet("Dashboard", 0)
                    ws.append(["Resumen", "Monto"])
                    ws.append(["Total Ingresos", 0])
                    ws.append(["Total Gastos", 0])
                    ws.append(["Total Ahorros", 0])
                    ws.append(["Balance Neto", 0])
                    needs_save = True
                    
                for sheet in ["Ingresos", "Gastos", "Ahorros"]:
                    if sheet not in wb.sheetnames:
                        ws = wb.create_sheet(sheet)
                        ws.append(self.columns)
                        needs_save = True
                        
                if needs_save:
                    wb.save(self.filename)
                    self._format_excel()
            except Exception as e:
                print(f"Error migrando Excel: {e}")

    def insert_record(self, record_type: str, category: str, description: str, amount: float) -> None:
        """Inserts a new record into the correct Excel sheet."""
        # El record_type que llega desde el form es 'Ingreso', 'Gasto', 'Ahorro'.
        # Aseguramos que el nombre de la hoja coincida copiando la 's' al final
        sheet_name = record_type + "s"
        if sheet_name not in self.sheets:
            sheet_name = "Ingresos"

        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nuevo_registro = pd.DataFrame([{
            "Fecha": fecha_actual,
            "Categoria": category,
            "Descripcion": description,
            "Monto": amount
        }])

        try:
            # Append a la hoja específica
            with pd.ExcelWriter(self.filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = writer.book
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # Write rows below the current max row
                    for row in dataframe_to_rows(nuevo_registro, index=False, header=False):
                        sheet.append(row)
            
            self._update_dashboard_sheet()
            self._format_excel()
        except Exception as e:
            print(f"Error escribiendo al archivo Excel: {e}")
            raise e

    def delete_record(self, tipo: str, fecha: str) -> bool:
        """Deletes a record matching Tipo and exact Fecha from the Excel file."""
        sheet_name = tipo + "s"
        if not os.path.exists(self.filename) or sheet_name not in self.sheets:
            return False

        try:
            wb = load_workbook(self.filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_to_delete = None
                
                # Buscar la fila a eliminar (omitir fila 1 de encabezados)
                for row in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=1).value
                    if cell_val is not None:
                        # Pandas lo lee como datetime string, openpyxl puede devolver datetime object o string
                        cell_str = cell_val if isinstance(cell_val, str) else cell_val.strftime("%Y-%m-%d %H:%M:%S")
                        
                        if cell_str == fecha:
                            row_to_delete = row
                            break
                            
                if row_to_delete:
                    ws.delete_rows(row_to_delete)
                    wb.save(self.filename)
                    self._update_dashboard_sheet()
                    self._format_excel()
                    return True
                    
            return False
        except Exception as e:
            print(f"Error borrando en Excel: {e}")
            return False

    def _format_excel(self) -> None:
        """Applies professional formatting to the Excel file."""
        if not os.path.exists(self.filename):
            return
            
        try:
            wb = load_workbook(self.filename)
            
            # --- Formato general ---
            header_font = Font(color="FFFFFF", bold=True, size=12, name="Segoe UI")
            header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color="E0E0E0"), 
                right=Side(style='thin', color="E0E0E0"), 
                top=Side(style='thin', color="E0E0E0"), 
                bottom=Side(style='thin', color="E0E0E0")
            )
            data_font = Font(size=11, name="Segoe UI")
            fill_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # --- Formatear Dashboard de Resumen ---
            if "Dashboard" in wb.sheetnames:
                ws_dash = wb["Dashboard"]
                for cell in ws_dash[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for row in ws_dash.iter_rows(min_row=2, max_row=5, min_col=1, max_col=2):
                    for cell in row:
                        cell.font = data_font
                        cell.border = thin_border
                        if cell.column_letter == 'B':
                            cell.number_format = '"$"#,##0.00'
                            cell.font = Font(size=11, name="Segoe UI", bold=True, color="1E3A8A")
                
                ws_dash.column_dimensions['A'].width = 25
                ws_dash.column_dimensions['B'].width = 20
            
            for sheet_name in ["Ingresos", "Gastos", "Ahorros"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # --- Formato de Encabezados ---
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # --- Formato de Filas de Datos ---
                    row_num = 1
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
                        row_num += 1
                        is_even = row_num % 2 == 0
                        current_fill = fill_light if is_even else fill_white
                        
                        for cell in row:
                            cell.font = data_font
                            cell.border = thin_border
                            cell.fill = current_fill
                            cell.alignment = Alignment(vertical="center")
                            
                            # Formato a la columna Fecha (A)
                            if cell.column_letter == 'A':
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            # Formato a la columna Monto (D)
                            elif cell.column_letter == 'D':
                                cell.number_format = '"$"#,##0.00'
                                # Monto en negrilla y ligeramente azul para resaltar
                                cell.font = Font(size=11, name="Segoe UI", bold=True, color="1E3A8A")

                    # --- Dimensiones de las Columnas ---
                    ws.column_dimensions['A'].width = 22 # Fecha
                    ws.column_dimensions['B'].width = 25 # Categoria
                    ws.column_dimensions['C'].width = 40 # Descripcion
                    ws.column_dimensions['D'].width = 18 # Monto
            
            wb.save(self.filename)
        except Exception as e:
            print(f"Error formateando Excel: {e}")
            
    def _update_dashboard_sheet(self) -> None:
        """Updates the totals in the Dashboard sheet in the Excel file."""
        if not os.path.exists(self.filename):
            return
            
        resumen = self.get_summary() # Totales históricos
        
        try:
            wb = load_workbook(self.filename)
            if "Dashboard" in wb.sheetnames:
                ws = wb["Dashboard"]
                ws["B2"] = resumen.get("Ingreso", 0)
                ws["B3"] = resumen.get("Gasto", 0)
                ws["B4"] = resumen.get("Ahorro", 0)
                ws["B5"] = resumen.get("Balance_Disponible", 0)
                wb.save(self.filename)
        except Exception as e:
            print(f"Error actualizando pestaña Dashboard de Excel: {e}")
            
    def get_summary(self, month: Optional[int] = None, year: Optional[int] = None) -> Dict[str, float]:
        """Returns the summary of Income, Expenses, and Savings for a given month/year."""
        resumen = {"Ingreso": 0.0, "Gasto": 0.0, "Ahorro": 0.0, "Balance_Disponible": 0.0}
        
        if not os.path.exists(self.filename):
            return resumen

        for sheet_name, tipo in [("Ingresos", "Ingreso"), ("Gastos", "Gasto"), ("Ahorros", "Ahorro")]:
            try:
                df = pd.read_excel(self.filename, sheet_name=sheet_name)
                if len(df) > 0:
                    df["Fecha"] = pd.to_datetime(df["Fecha"])
                    
                    if month and year:
                        df = df[(df["Fecha"].dt.month == month) & (df["Fecha"].dt.year == year)]
                    elif month:
                        df = df[df["Fecha"].dt.month == month]
                    elif year:
                        df = df[df["Fecha"].dt.year == year]
                        
                    resumen[tipo] = float(df["Monto"].sum())
            except ValueError:
                # La hoja podría no existir
                pass

        # Calculate logical balance
        resumen["Balance_Disponible"] = resumen["Ingreso"] - resumen["Gasto"]
        
        return resumen
    
    def get_recent_records(self, limit: int = 10) -> List[Dict[str, Any]]:
        if not os.path.exists(self.filename):
            return []
            
        dfs = []
        for sheet_name, tipo in [("Ingresos", "Ingreso"), ("Gastos", "Gasto"), ("Ahorros", "Ahorro")]:
            try:
                df = pd.read_excel(self.filename, sheet_name=sheet_name)
                if len(df) > 0:
                    df["Tipo"] = tipo
                    dfs.append(df)
            except ValueError:
                pass
                
        if not dfs:
            return []
            
        df_all = pd.concat(dfs, ignore_index=True)
        df_all["Fecha"] = pd.to_datetime(df_all["Fecha"])
        df_all = df_all.sort_values(by="Fecha", ascending=False).head(limit)
        
        # Convert date to string for JSON serialization
        df_all["Fecha"] = df_all["Fecha"].dt.strftime("%Y-%m-%d %H:%M:%S")
        
        return df_all.to_dict(orient="records")

# Instancia global de la BD
db = ExcelDB()
